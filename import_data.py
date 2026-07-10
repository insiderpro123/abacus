"""
One-time importer: load the current Excel files + project_icons.json into the database.

Usage:
    python import_data.py --yes            # wipe reference+data and reload from Excel

Reads (from this folder):
    Process_Abacus_Vertical.xlsm  (sheet WP_Master)  - work packages + RAG status
    Abacus_prompts into outcomes_20 01 26.xlsx (Sheet1) - guidance per point code
    project_icons.json                                - per-wp emoji (optional)

Writes into the DB configured by models.py (SQLite locally, Postgres via DATABASE_URL).
The parsing mirrors the old app.py Excel logic so nothing changes meaning.
"""

import json
import os
import re
import shutil
import sys
import tempfile

import openpyxl

from models import (
    HERE, init_db, SessionLocal,
    Process, Subprocess, WorkPackage, WpStatus, WpFinished,
)

WORKBOOK = os.path.join(HERE, "Process_Abacus_Vertical.xlsm")
PROMPTS = os.path.join(HERE, "Abacus_prompts into outcomes_20 01 26.xlsx")
ICONS_FILE = os.path.join(HERE, "project_icons.json")
SHEET = "WP_Master"
PROMPTS_SHEET = "Sheet1"

GROUP_ROW, HEADER_ROW, FIRST_DATA_ROW = 4, 6, 7
COL_WP_ID, COL_CLIENT, COL_NAME, COL_STATUS, COL_POINTS = 0, 1, 2, 3, 4


def _clean(v):
    return "" if v is None else str(v).strip()


def _load_locked_safe(path):
    try:
        return openpyxl.load_workbook(path, data_only=True, read_only=True, keep_vba=False)
    except PermissionError:
        tmp = os.path.join(tempfile.gettempdir(), "_abacus_import_" + os.path.basename(path))
        shutil.copy2(path, tmp)
        return openpyxl.load_workbook(tmp, data_only=True, read_only=True, keep_vba=False)


def _parse_group_header(text):
    m = re.match(r"\s*(\d+)\.?\s*(.*)", text)
    return (int(m.group(1)), m.group(2).strip()) if m else (None, text)


def _sub_code(text, group_num, fallback_idx):
    m = re.match(r"\s*(\d+\.\d+)", text)
    if m:
        return m.group(1)
    m = re.match(r"\s*(\d+)\.", text)
    if m:
        return f"{m.group(1)}.{fallback_idx}"
    return f"{group_num}.{fallback_idx}"


def build_structure(rows):
    group_row, header_row = rows[GROUP_ROW - 1], rows[HEADER_ROW - 1]
    groups, current = [], None
    for i in range(COL_POINTS + 1, len(header_row)):
        g = _clean(group_row[i]) if i < len(group_row) else ""
        sub = _clean(header_row[i])
        if g and g[0].isdigit():
            num, title = _parse_group_header(g)
            current = {"num": num, "title": title, "subs": [], "finished_col": None}
            groups.append(current)
        if current is None:
            continue
        if sub.lower() == "finished?":
            current["finished_col"] = i
        elif sub:
            idx = len(current["subs"]) + 1
            code = _sub_code(sub, current["num"], idx)
            label = re.sub(r"^\s*\d+(\.\d+)?\.?\s*", "", sub).strip()
            current["subs"].append({"col": i, "code": code, "label": label})
    return groups


def _prompts_code(qtext):
    m = re.match(r"\s*(\d+\.\d+)", qtext)
    if m:
        return m.group(1)
    m = re.match(r"\s*(\d+)\.", qtext)
    return f"{m.group(1)}.1" if m else None


def get_reference():
    try:
        ws = _load_locked_safe(PROMPTS)[PROMPTS_SHEET]
    except Exception as e:
        print("WARNING: could not read prompts file:", e)
        return {}
    ref = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        q = _clean(row[0]) if row else ""
        if not q or q.lower().startswith("finished"):
            continue
        code = _prompts_code(q)
        if not code or code in ref:
            continue

        def col(i):
            return _clean(row[i]) if len(row) > i else ""

        comment = col(5)
        extra = col(6)
        if extra:
            comment = (comment + "\n" + extra).strip()
        ref[code] = {
            "outcomes": col(1), "operational": col(2), "top_level": col(3),
            "assets": col(4), "comment": comment,
        }
    return ref


def _norm_value(raw):
    """Excel cell -> stored value ('1'/'2'/'3'/'N/R') or None if not started."""
    s = _clean(raw)
    if s in ("1", "2", "3"):
        return s
    if s.upper() in ("N/R", "NR", "N/A", "NOT REQUIRED"):
        return "N/R"
    return None


def main():
    if "--yes" not in sys.argv:
        print("This wipes the reference + work-package tables and reloads from Excel.")
        print("Re-run with:  python import_data.py --yes")
        return

    init_db()
    ws = _load_locked_safe(WORKBOOK)[SHEET]
    rows = list(ws.iter_rows(values_only=True))
    groups = build_structure(rows)
    ref = get_reference()

    try:
        icons = json.load(open(ICONS_FILE, encoding="utf-8"))
    except Exception:
        icons = {}

    s = SessionLocal()
    try:
        # wipe (order respects FKs)
        for model in (WpStatus, WpFinished, WorkPackage, Subprocess, Process):
            s.query(model).delete()
        s.flush()

        # reference: processes + subprocesses
        code_to_process = {}
        for g in groups:
            s.add(Process(num=g["num"], title=g["title"]))
            for seq, sub in enumerate(g["subs"], start=1):
                code = sub["code"]
                code_to_process[code] = g["num"]
                r = ref.get(code, {})
                s.add(Subprocess(
                    code=code, process_num=g["num"], seq=seq,
                    question=sub["label"],
                    outcomes=r.get("outcomes", ""), operational=r.get("operational", ""),
                    top_level=r.get("top_level", ""), assets=r.get("assets", ""),
                    comment=r.get("comment", ""),
                ))
        s.flush()

        # work packages + per-point status + finished flags
        n_wp = n_status = n_fin = 0
        for r in rows[FIRST_DATA_ROW - 1:]:
            if not r or _clean(r[COL_NAME]) == "" or _clean(r[COL_WP_ID]) == "":
                continue
            try:
                wid = int(float(_clean(r[COL_WP_ID])))
            except ValueError:
                continue
            s.add(WorkPackage(
                id=wid, client=_clean(r[COL_CLIENT]), name=_clean(r[COL_NAME]),
                status=_clean(r[COL_STATUS]) or "Active",
                points=_clean(r[COL_POINTS]), icon=_clean(icons.get(str(wid), "")),
            ))
            n_wp += 1
            for g in groups:
                for sub in g["subs"]:
                    val = _norm_value(r[sub["col"]]) if sub["col"] < len(r) else None
                    if val is not None:
                        s.add(WpStatus(wp_id=wid, code=sub["code"], value=val))
                        n_status += 1
                fc = g["finished_col"]
                if fc is not None and fc < len(r) and _clean(r[fc]).lower() == "x":
                    s.add(WpFinished(wp_id=wid, process_num=g["num"]))
                    n_fin += 1

        s.commit()
        print(f"Imported: {len(groups)} processes, "
              f"{sum(len(g['subs']) for g in groups)} sub-points, "
              f"{n_wp} work packages, {n_status} status cells, {n_fin} finished flags.")
    except Exception:
        s.rollback()
        raise
    finally:
        s.close()


if __name__ == "__main__":
    main()
